import openpyxl
import logging
from openpyxl.styles import Font, Alignment

logger = logging.getLogger(__name__)

def write_to_excel(user_data, field_list, output_file):
    """
    Записывает данные в Excel-файл в формате столбцов по позициям полей.
    user_data: словарь {telegram_id: {position: value}}
    field_list: список полей [{'id': id, 'position': pos, 'label': label}]
    output_file: путь к выходному файлу
    """
    try:
        # Создаем новую книгу Excel
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Responses"

        # Первая колонка - Telegram ID
        ws.cell(row=1, column=1, value="Telegram ID").font = Font(bold=True)

        # Создаем заголовки для каждого поля по позициям
        max_position = max(field['position'] for field in field_list) if field_list else 0
        position_to_col = {field['position']: col_num for col_num, field in enumerate(field_list, 2)}

        for field in field_list:
            col_num = position_to_col[field['position']]
            cell = ws.cell(row=1, column=col_num, value=field['label'])
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')

        # Записываем данные пользователей
        row_num = 2
        for telegram_id, responses in user_data.items():
            ws.cell(row=row_num, column=1, value=telegram_id)

            # Заполняем ответы по позициям
            for position, value in responses.items():
                if position in position_to_col:
                    col_num = position_to_col[position]
                    ws.cell(row=row_num, column=col_num, value=value)

            row_num += 1

        # Автоматическая ширина столбцов
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)  # Ограничение максимальной ширины
            ws.column_dimensions[column].width = adjusted_width

        # Сохраняем файл
        wb.save(output_file)
        logger.info(f"Данные успешно записаны в файл {output_file}.")
    except Exception as e:
        logger.error(f"Ошибка при записи в Excel: {e}")
        raise